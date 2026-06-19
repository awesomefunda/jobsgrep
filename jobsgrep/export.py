"""Segmented job-sheet exporter.

Turns the cached job corpus into a single Excel workbook that's built to be fed
straight into an LLM with the user's resume:

  - "Start Here" sheet: ready-to-paste ranking prompts + per-family counts
  - one sheet per role family (Software Engineering, Data & ML, ...), each with a
    Level column so users can filter by seniority in Excel before ranking
  - an "All Jobs" master sheet

There is no scoring here — ranking happens on the user's machine with their own
LLM. This is pure data shaping.
"""
from __future__ import annotations

import logging
from collections import defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import RawJob
from .taxonomy import LEVEL_ORDER, ROLE_ORDER, classify_level, classify_role_family

logger = logging.getLogger("jobsgrep.export")

_HDR_FILL = PatternFill("solid", fgColor="1B2A4A")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_TITLE_FILL = PatternFill("solid", fgColor="0D5E4A")
_ACCENT_FILL = PatternFill("solid", fgColor="E8F5E9")
_THIN = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="top")
_TOP = Alignment(vertical="top")

_COLUMNS = [
    ("Company", 22), ("Title", 34), ("Level", 11), ("Location", 22),
    ("Remote?", 9), ("Salary", 16), ("Date Posted", 12), ("Source", 15),
    ("Job URL", 40), ("Description", 60),
]
_URL_COL = 9
_LEVEL_RANK = {lvl: i for i, lvl in enumerate(LEVEL_ORDER)}

_RANKING_PROMPT = (
    "You are my job-search assistant. Below (or in the attached spreadsheet) is a "
    "list of open jobs with columns: Company, Title, Level, Location, Remote, "
    "Salary, Posted, Source, URL. My resume follows at the bottom.\n\n"
    "1. Score every job 0-100 for fit with my resume and preferences.\n"
    "2. Return a ranked table: Rank | Company | Title | Level | Location | Score | "
    "one-line reason.\n"
    "3. For the top 10, add: why I'm a fit, my biggest gap for that role, and a "
    "2-sentence tailored application hook.\n"
    "4. Flag any that look like a stretch or likely need a referral.\n\n"
    "MY PREFERENCES (edit me): remote? = ; locations = ; target level = ; "
    "must-have tech = ; deal-breakers = .\n\n"
    "MY RESUME (paste below):\n[PASTE YOUR RESUME HERE]"
)

_EXTRA_PROMPTS = [
    ("Cover letter",
     "Using my resume and this job (Company / Title / URL from the row), write a "
     "concise 3-paragraph cover letter. Lead with the strongest overlap. No clichés."),
    ("Resume tailoring",
     "Compare my resume to this role. Give: 3 bullets to rewrite to match their "
     "language, keywords I'm missing, and one line to cut. Quote exact lines."),
    ("Skill-gap check",
     "Given this role and my resume: the 3 most important skills I lack, a <2-week "
     "project to demonstrate each, and an honest 'apply now vs upskill first' call."),
]


def _salary(job: RawJob) -> str:
    if job.salary_text:
        return job.salary_text
    if job.salary_min or job.salary_max:
        lo = f"{int(job.salary_min):,}" if job.salary_min else "?"
        hi = f"{int(job.salary_max):,}" if job.salary_max else "?"
        return f"{lo}–{hi}"
    return ""


def _group_by_family(jobs: list[RawJob]) -> dict[str, list[RawJob]]:
    groups: dict[str, list[RawJob]] = defaultdict(list)
    for job in jobs:
        groups[classify_role_family(job.title)].append(job)
    # Sort each group by seniority then company for easy scanning.
    for fam in groups:
        groups[fam].sort(
            key=lambda j: (_LEVEL_RANK.get(classify_level(j.title), 99), j.company.lower())
        )
    return groups


def _write_jobs_sheet(ws, jobs: list[RawJob]) -> None:
    for col, (name, width) in enumerate(_COLUMNS, 1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = _HDR_FILL
        c.font = _HDR_FONT
        c.alignment = _WRAP
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COLUMNS))}1"

    for r, job in enumerate(jobs, 2):
        desc = (job.description or "").strip().replace("\n", " ")
        row = [
            job.company,
            job.title,
            classify_level(job.title),
            job.location,
            "Yes" if job.remote else "No",
            _salary(job),
            job.date_posted,
            job.source,
            job.url,
            desc[:500] + ("…" if len(desc) > 500 else ""),
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _WRAP if col in (2, 10) else _TOP
        url_cell = ws.cell(row=r, column=_URL_COL)
        if job.url:
            url_cell.hyperlink = job.url
            url_cell.font = Font(color="0563C1", underline="single")

    ws.row_dimensions[1].height = 26


def _write_start_here(ws, groups: dict[str, list[RawJob]], total: int) -> None:
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 95

    def section(row: int, text: str, fill: PatternFill) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill
        c.font = _HDR_FONT
        c.alignment = _WRAP
        c.border = _BORDER

    section(1, "JOBSGREP — OPEN JOBS PACK", _TITLE_FILL)
    ws.cell(row=2, column=1, value="Generated (UTC)").font = Font(bold=True)
    ws.cell(row=2, column=2, value=datetime.utcnow().strftime("%Y-%m-%d %H:%M"))
    ws.cell(row=3, column=1, value="Total jobs").font = Font(bold=True)
    ws.cell(row=3, column=2, value=total)

    section(5, "HOW TO USE", _HDR_FILL)
    howto = (
        "1. Pick the role tab that matches you (Software Engineering, Data & ML, …) "
        "or use the All Jobs tab.\n"
        "2. Filter by the Level column in Excel to your seniority.\n"
        "3. Copy the rows you care about, paste them into Claude/ChatGPT with your "
        "resume, and use the ranking prompt below to get a personalized shortlist."
    )
    ws.merge_cells(start_row=6, start_column=1, end_row=6, end_column=2)
    hc = ws.cell(row=6, column=1, value=howto)
    hc.alignment = _WRAP
    hc.border = _BORDER
    hc.fill = _ACCENT_FILL
    ws.row_dimensions[6].height = 70

    section(8, "RANKING PROMPT  (copy → paste into your LLM with the job rows + your resume)", _TITLE_FILL)
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=2)
    pc = ws.cell(row=9, column=1, value=_RANKING_PROMPT)
    pc.alignment = _WRAP
    pc.border = _BORDER
    pc.fill = _ACCENT_FILL
    ws.row_dimensions[9].height = 230

    section(11, "MORE PROMPTS", _HDR_FILL)
    r = 12
    for name, prompt in _EXTRA_PROMPTS:
        ws.cell(row=r, column=1, value=name).font = Font(bold=True)
        nc = ws.cell(row=r, column=2, value=prompt)
        nc.alignment = _WRAP
        nc.border = _BORDER
        ws.row_dimensions[r].height = 46
        r += 1

    r += 1
    section(r, "WHAT'S IN THIS FILE", _HDR_FILL)
    r += 1
    ws.cell(row=r, column=1, value="Tab").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Jobs").font = Font(bold=True)
    r += 1
    for fam in ROLE_ORDER:
        if groups.get(fam):
            ws.cell(row=r, column=1, value=fam)
            ws.cell(row=r, column=2, value=len(groups[fam]))
            r += 1


def _safe_sheet_name(name: str, used: set[str]) -> str:
    # Excel: max 31 chars, no : \ / ? * [ ]
    clean = name
    for ch in r':\/?*[]':
        clean = clean.replace(ch, "-")
    clean = clean[:31] or "Sheet"
    base, i = clean, 1
    while clean.lower() in used:
        suffix = f" {i}"
        clean = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(clean.lower())
    return clean


def export_segmented(jobs: list[RawJob], output_dir: Path) -> Path:
    """Build the full segmented workbook (all role tabs + All Jobs) and return its path."""
    from .taxonomy import filter_tech
    jobs = filter_tech(jobs)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"jobsgrep_jobs_{ts}.xlsx"

    groups = _group_by_family(jobs)

    wb = Workbook()
    used: set[str] = set()

    start = wb.active
    start.title = _safe_sheet_name("Start Here", used)
    _write_start_here(start, groups, len(jobs))

    # One sheet per non-empty role family, in display order.
    for fam in ROLE_ORDER:
        fam_jobs = groups.get(fam)
        if not fam_jobs:
            continue
        ws = wb.create_sheet(_safe_sheet_name(fam, used))
        _write_jobs_sheet(ws, fam_jobs)

    # Master sheet with everything.
    all_ws = wb.create_sheet(_safe_sheet_name("All Jobs", used))
    _write_jobs_sheet(all_ws, jobs)

    wb.save(str(out_path))
    logger.info("exported %d jobs across %d role tabs -> %s",
                len(jobs), sum(1 for f in ROLE_ORDER if groups.get(f)), out_path.name)
    return out_path


def build_family_workbook(jobs: list[RawJob], family: str, output_dir: Path) -> Path | None:
    """Build a one-family workbook (Start Here + that role tab). Returns path or None if empty."""
    from .insights import ROLE_SLUGS
    from .taxonomy import filter_tech

    fam_jobs = [j for j in filter_tech(jobs) if classify_role_family(j.title) == family]
    if not fam_jobs:
        return None
    fam_jobs.sort(
        key=lambda j: (_LEVEL_RANK.get(classify_level(j.title), 99), j.company.lower())
    )

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    slug = ROLE_SLUGS.get(family, "jobs")
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    out_path = output_dir / f"jobsgrep_{slug}_{ts}.xlsx"

    wb = Workbook()
    used: set[str] = set()
    start = wb.active
    start.title = _safe_sheet_name("Start Here", used)
    _write_start_here(start, {family: fam_jobs}, len(fam_jobs))
    ws = wb.create_sheet(_safe_sheet_name(family, used))
    _write_jobs_sheet(ws, fam_jobs)

    wb.save(str(out_path))
    logger.info("exported %d %s jobs -> %s", len(fam_jobs), family, out_path.name)
    return out_path
