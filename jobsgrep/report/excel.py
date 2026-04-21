"""Excel report generator using openpyxl — 3 sheets."""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ..models import ScoredJob, SearchTask

logger = logging.getLogger("jobsgrep.report")

# ─── Styles ──────────────────────────────────────────────────────────────────
_HDR_JOB   = PatternFill("solid", fgColor="1B2A4A")   # dark navy  — auto-filled cols
_HDR_TRACK = PatternFill("solid", fgColor="0D5E4A")   # dark teal  — user tracking cols
_HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)

_GREEN_FILL  = PatternFill("solid", fgColor="E8F5E9")  # score ≥ 0.9
_YELLOW_FILL = PatternFill("solid", fgColor="FFFDE7")  # score 0.8–0.9
_WHITE_FILL  = PatternFill("solid", fgColor="FFFFFF")
_TRACK_FILL  = PatternFill("solid", fgColor="F0FAF7")  # faint teal tint on tracking cols

_THIN   = Side(style="thin", color="D0D0D0")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP   = Alignment(wrap_text=True, vertical="top")
_TOP    = Alignment(vertical="top")
_CENTER = Alignment(horizontal="center", vertical="top")

# Tracking column Status dropdown choices
_STATUS_OPTIONS = "New,Saved,Applied,Phone Screen,Technical,Onsite,Offer,Rejected,Withdrawn"

# ─── Column layout ────────────────────────────────────────────────────────────
#
#  ← JobsGrep auto-filled (cols 1–15) ─────────────────────────────────────────┐  ← Your tracking (cols 16–23) ──┐
#  Rank Score Company Title RoleType Seniority Location Remote Salary URL Date  │  Status Applied Outreach …      │
#  Source Matching Missing RedFlags                                              │                                 │
#
_JOB_HEADERS = [
    "Rank", "Score", "Company", "Title", "Role Type", "Seniority", "Location", "Remote?",
    "Salary", "Job URL", "Date Posted", "Source",
    "Matching Skills", "Missing Skills", "Red Flags",
]
_TRACK_HEADERS = [
    "Status", "Applied Date", "Outreach Sent", "Response?",
    "Phone Screen", "Interview", "Offer", "Notes",
]
_ALL_HEADERS  = _JOB_HEADERS + _TRACK_HEADERS
_TRACK_START  = len(_JOB_HEADERS) + 1   # first tracking column index (1-based)
_URL_COL      = 10                       # Job URL column index (after adding Role Type + Seniority)


def _hdr(ws, values: list[str]) -> None:
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = _HDR_TRACK if col >= _TRACK_START else _HDR_JOB
        cell.font = _HDR_FONT
        cell.alignment = _WRAP
        cell.border = _BORDER


def _score_fill(score: float) -> PatternFill:
    if score >= 0.9:
        return _GREEN_FILL
    if score >= 0.8:
        return _YELLOW_FILL
    return _WHITE_FILL


def _col_widths(ws, widths: dict[int, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def generate_report(
    scored_jobs: list[ScoredJob],
    task: SearchTask,
    output_dir: Path,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    task_suffix = f"_{task.task_id}" if task.task_id else ""
    out_path = output_dir / f"jobsgrep_report_{ts}{task_suffix}.xlsx"

    wb = Workbook()
    _sheet_tracker(wb, scored_jobs)
    _sheet_all_jobs(wb, scored_jobs)
    _sheet_ai_toolkit(wb, scored_jobs, task)
    _sheet_summary(wb, task)

    wb.save(str(out_path))
    logger.info("report saved: %s (%d matched jobs)", out_path.name, len(scored_jobs))
    return out_path


def _sheet_tracker(wb: Workbook, scored_jobs: list[ScoredJob]) -> None:
    """Sheet 1: Job Tracker — auto-filled job data + blank tracking columns."""
    ws = wb.active
    ws.title = "Job Tracker"

    _hdr(ws, _ALL_HEADERS)

    # Freeze: keep rank+score+company+title visible while scrolling right,
    # and header row visible while scrolling down
    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_ALL_HEADERS))}1"

    # Status dropdown validation applied to the whole Status column
    last_row = max(len(scored_jobs) + 1, 2)
    status_col_letter = get_column_letter(_TRACK_START)
    dv = DataValidation(
        type="list",
        formula1=f'"{_STATUS_OPTIONS}"',
        allow_blank=True,
        showErrorMessage=False,
    )
    dv.sqref = f"{status_col_letter}2:{status_col_letter}{last_row}"
    ws.add_data_validation(dv)

    for rank, sj in enumerate(scored_jobs, 1):
        job   = sj.job
        score = sj.score
        r     = rank + 1
        job_fill   = _score_fill(score.fit_score)

        row_data = [
            rank,
            round(score.fit_score, 2),
            job.company,
            job.title,
            score.role_type,
            score.seniority_level,
            job.location,
            "Yes" if job.remote else "No",
            score.salary_range or job.salary_text or "",
            job.url,
            job.date_posted,
            job.source,
            ", ".join(score.matching_skills),
            ", ".join(score.missing_skills),
            ", ".join(score.red_flags),
            # Tracking cols — pre-fill Status as "New"
            "New", "", "", "", "", "", "", "",
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            if col < _TRACK_START:
                cell.fill  = job_fill
                cell.alignment = _WRAP
            else:
                cell.fill  = _TRACK_FILL
                cell.alignment = _TOP

        # Clickable URL
        url_cell = ws.cell(row=r, column=_URL_COL)
        if job.url:
            url_cell.hyperlink = job.url
            url_cell.font = Font(color="0563C1", underline="single")

        # Score: bold + coloured font
        score_cell = ws.cell(row=r, column=2)
        if score.fit_score >= 0.9:
            score_cell.font = Font(bold=True, color="2E7D32")
        elif score.fit_score >= 0.8:
            score_cell.font = Font(bold=True, color="F57F17")
        else:
            score_cell.font = Font(bold=True, color="555555")

    # Column widths
    _col_widths(ws, {
        1:  6,   # Rank
        2:  7,   # Score
        3: 22,   # Company
        4: 30,   # Title
        5: 18,   # Role Type
        6: 11,   # Seniority
        7: 18,   # Location
        8:  8,   # Remote?
        9: 16,   # Salary
        10: 38,  # Job URL
        11: 12,  # Date Posted
        12: 16,  # Source
        13: 28,  # Matching Skills
        14: 24,  # Missing Skills
        15: 24,  # Red Flags
        # Tracking
        16: 14,  # Status
        17: 13,  # Applied Date
        18: 14,  # Outreach Sent
        19: 12,  # Response?
        20: 14,  # Phone Screen
        21: 13,  # Interview
        22: 12,  # Offer
        23: 30,  # Notes
    })

    ws.row_dimensions[1].height = 32
    for r in range(2, len(scored_jobs) + 2):
        ws.row_dimensions[r].height = 48

    # Visual separator: thicker left border on first tracking column
    thick = Side(style="medium", color="0D5E4A")
    for r in range(1, len(scored_jobs) + 2):
        cell = ws.cell(row=r, column=_TRACK_START)
        cell.border = Border(
            left=thick,
            right=_THIN,
            top=_THIN,
            bottom=_THIN,
        )

    # Instruction row note in cell A1 comment area — use a helper cell below data
    note_row = len(scored_jobs) + 3
    note_cell = ws.cell(row=note_row, column=_TRACK_START,
                        value="← Fill in the teal columns as you progress through your job search")
    note_cell.font = Font(italic=True, color="888888", size=9)


def _sheet_all_jobs(wb: Workbook, scored_jobs: list[ScoredJob]) -> None:
    """Sheet 2: All Jobs Found — reference list with scores."""
    ws = wb.create_sheet("All Jobs Found")

    headers = ["Company", "Title", "Location", "Remote?", "Score", "Source", "Job URL", "Date Posted"]
    for col, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.fill = _HDR_JOB
        cell.font = _HDR_FONT
        cell.alignment = _WRAP
        cell.border = _BORDER

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    for r, sj in enumerate(scored_jobs, 2):
        job = sj.job
        row_data = [
            job.company, job.title, job.location,
            "Yes" if job.remote else "No",
            round(sj.score.fit_score, 2),
            job.source, job.url, job.date_posted,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border = _BORDER
            cell.alignment = _WRAP

        url_cell = ws.cell(row=r, column=7)
        if job.url:
            url_cell.hyperlink = job.url
            url_cell.font = Font(color="0563C1", underline="single")

    _col_widths(ws, {1: 22, 2: 32, 3: 20, 4: 8, 5: 7, 6: 18, 7: 38, 8: 12})
    ws.row_dimensions[1].height = 28
    for r in range(2, len(scored_jobs) + 2):
        ws.row_dimensions[r].height = 30


def _sheet_ai_toolkit(wb: Workbook, scored_jobs: list[ScoredJob], task: SearchTask) -> None:
    """Sheet 3: AI Toolkit — resume paste area + per-job LLM prompts.

    Upload this sheet (or the whole file) to Claude / ChatGPT / Codex to:
      • Write personalized cover letters
      • Tailor your resume for each role
      • Draft LinkedIn outreach messages
      • Identify skill gaps and a learning plan
    """
    ws = wb.create_sheet("AI Toolkit")

    BLUE   = PatternFill("solid", fgColor="1B2A4A")
    TEAL   = PatternFill("solid", fgColor="0D5E4A")
    AMBER  = PatternFill("solid", fgColor="FF8F00")
    LGRAY  = PatternFill("solid", fgColor="F5F5F5")
    LGREEN = PatternFill("solid", fgColor="E8F5E9")
    WHITE  = PatternFill("solid", fgColor="FFFFFF")
    WHT_FN = Font(color="FFFFFF", bold=True, size=10)
    BLD    = Font(bold=True)
    ITL    = Font(italic=True, color="555555", size=9)

    def _section(row: int, text: str, fill: PatternFill, cols: int = 2) -> None:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
        c = ws.cell(row=row, column=1, value=text)
        c.fill = fill; c.font = WHT_FN; c.alignment = _WRAP; c.border = _BORDER

    def _label(row: int, col: int, text: str) -> None:
        c = ws.cell(row=row, column=col, value=text)
        c.font = BLD; c.fill = LGRAY; c.border = _BORDER; c.alignment = _TOP

    def _cell(row: int, col: int, text: str, fill=WHITE) -> None:
        c = ws.cell(row=row, column=col, value=text)
        c.fill = fill; c.border = _BORDER; c.alignment = _WRAP

    # ── Section 1: How to use ────────────────────────────────────────────────
    _section(1, "HOW TO USE THIS SHEET", BLUE, cols=2)
    instructions = (
        "1. PASTE YOUR RESUME in the yellow box below (plain text is fine).\n"
        "2. UPLOAD THIS FILE to Claude (claude.ai), ChatGPT, or any LLM tool — "
        "or copy individual prompts from the table at the bottom.\n"
        "3. The LLM will use your resume + the job data to write cover letters, "
        "tailor your resume, draft outreach messages, and identify skill gaps.\n\n"
        "TIP: Tell the LLM  →  \"Use my resume and the job data in the AI Toolkit "
        "sheet to help me apply to the top 5 jobs. Start with cover letters.\""
    )
    ws.merge_cells(start_row=2, start_column=1, end_row=5, end_column=2)
    ic = ws.cell(row=2, column=1, value=instructions)
    ic.fill = LGREEN; ic.alignment = _WRAP; ic.border = _BORDER
    ws.row_dimensions[2].height = 80

    # ── Section 2: Resume paste area ────────────────────────────────────────
    _section(6, "YOUR RESUME  (paste plain text below — used by the LLM for all prompts)", AMBER, cols=2)
    ws.merge_cells(start_row=7, start_column=1, end_row=22, end_column=2)
    rc = ws.cell(row=7, column=1,
                 value="← Paste your resume here. Include: name, contact, summary, "
                       "work experience (company / title / dates / bullets), "
                       "education, and skills.")
    rc.font = Font(italic=True, color="AAAAAA")
    rc.fill = PatternFill("solid", fgColor="FFFDE7")
    rc.alignment = _WRAP; rc.border = _BORDER
    ws.row_dimensions[7].height = 200

    # ── Section 3: Prompt templates ──────────────────────────────────────────
    _section(23, "READY-TO-USE PROMPT TEMPLATES  (copy & paste into any LLM)", TEAL, cols=2)
    templates = [
        ("Cover Letter",
         "I'm applying to [COMPANY] for the [TITLE] role (URL: [URL]).\n"
         "Using my resume above and the job description below, write a concise, "
         "compelling cover letter (3 paragraphs). Lead with the strongest overlap "
         "between my experience and what they need. Avoid clichés.\n\n"
         "Job description: [PASTE JOB DESCRIPTION]"),
        ("Resume Tailoring",
         "Compare my resume to the [TITLE] role at [COMPANY].\n"
         "List: (1) 3 bullet points I should rewrite to better match their language, "
         "(2) any keywords from the JD I'm missing entirely, "
         "(3) one line I should remove because it's irrelevant.\n"
         "Be specific — quote the exact resume lines and suggest rewrites."),
        ("LinkedIn Outreach",
         "Write a short LinkedIn message (< 75 words) to the hiring manager at "
         "[COMPANY] for the [TITLE] role. I want to express genuine interest, "
         "mention one specific thing about the company that resonates with my "
         "background, and ask for a 15-min call. Keep it warm but professional."),
        ("Skill Gap Analysis",
         "Based on the [TITLE] job at [COMPANY] and my resume, tell me:\n"
         "(1) The 3 most important skills I'm missing\n"
         "(2) For each missing skill: a free resource or project I can do in < 2 weeks "
         "to demonstrate it\n"
         "(3) An honest assessment — should I apply now or upskill first?"),
        ("Batch Score All Jobs",
         "I have a list of job listings below (from the Job Tracker sheet). "
         "Score each one from 0–10 based on fit with my resume. "
         "Output a ranked table: Rank | Company | Title | Score | One-line reason.\n"
         "Prioritise roles where my experience directly matches ≥ 70% of requirements."),
    ]
    for i, (name, prompt) in enumerate(templates):
        r = 24 + i * 3
        _label(r, 1, name)
        ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=2)
        pc = ws.cell(row=r, column=2, value=prompt)
        pc.fill = LGREEN; pc.alignment = _WRAP; pc.border = _BORDER
        ws.row_dimensions[r].height = 60
        ws.row_dimensions[r + 1].height = 1
        ws.row_dimensions[r + 2].height = 1

    # ── Section 4: Per-job data table ────────────────────────────────────────
    table_start = 24 + len(templates) * 3 + 1
    _section(table_start, "JOB DATA  (company, description, skills — used by the LLM prompts above)", BLUE, cols=6)

    col_headers = ["#", "Company", "Title", "Location", "Required Skills / Missing Skills", "Job Description (first 1500 chars)"]
    col_widths   = [4,   20,        28,      18,          38,                                  70]
    for col, (hdr, w) in enumerate(zip(col_headers, col_widths), 1):
        c = ws.cell(row=table_start + 1, column=col, value=hdr)
        c.fill = BLUE; c.font = WHT_FN; c.alignment = _WRAP; c.border = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = w

    for rank, sj in enumerate(scored_jobs, 1):
        job   = sj.job
        score = sj.score
        r     = table_start + 1 + rank

        skills_text = ""
        if score.matching_skills:
            skills_text += "Have: " + ", ".join(score.matching_skills)
        if score.missing_skills:
            skills_text += ("\n" if skills_text else "") + "Missing: " + ", ".join(score.missing_skills)
        if not skills_text:
            skills_text = job.source  # fallback if no LLM scoring

        desc = (job.description or "").strip()
        desc_preview = desc[:1500] + ("…" if len(desc) > 1500 else "")

        row_vals = [rank, job.company, job.title, job.location, skills_text, desc_preview]
        fill = _score_fill(score.fit_score)
        for col, val in enumerate(row_vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill if col <= 4 else WHITE
            c.border = _BORDER; c.alignment = _WRAP
        ws.row_dimensions[r].height = 72

    # Column widths for resume + instructions columns (A, B)
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 80

    # Row heights for fixed sections
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[6].height = 22
    ws.row_dimensions[23].height = 22
    for ri in range(7, 23):
        ws.row_dimensions[ri].height = 14


def _sheet_summary(wb: Workbook, task: SearchTask) -> None:
    """Sheet 3: Search Summary."""
    ws = wb.create_sheet("Search Summary")

    from ..config import get_settings
    settings = get_settings()

    lbl_fill = PatternFill("solid", fgColor="E8EAF6")
    lbl_font = Font(bold=True)

    def _row(r: int, label: str, value: str) -> None:
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = lbl_font
        lc.fill = lbl_fill
        lc.border = _BORDER
        vc = ws.cell(row=r, column=2, value=value)
        vc.alignment = _WRAP
        vc.border = _BORDER

    rows = [
        ("Original Query",              task.query),
        ("Mode",                        settings.jobsgrep_mode.value),
        ("Generated At",                task.created_at.strftime("%Y-%m-%d %H:%M UTC") if task.created_at else ""),
        ("Completed At",                task.completed_at.strftime("%Y-%m-%d %H:%M UTC") if task.completed_at else ""),
        ("Total Jobs Found",            str(task.total_jobs_found)),
        ("Jobs in Tracker (≥ threshold)", str(task.total_jobs_scored)),
        ("Min Fit Score",               str(settings.min_fit_score)),
        ("Sources Searched",            ", ".join(task.sources_searched)),
        ("",                            ""),
        ("Jobs Per Source",             ""),
    ]
    for r, (label, value) in enumerate(rows, 1):
        _row(r, label, value)

    r = len(rows) + 1
    for source, count in task.jobs_per_source.items():
        ws.cell(row=r, column=1, value=f"  {source}").font = Font(italic=True)
        ws.cell(row=r, column=2, value=count)
        r += 1

    if task.parsed_query:
        ws.cell(row=r, column=1, value="Parsed Query Interpretation").font = lbl_font
        r += 1
        pq = task.parsed_query
        for label, value in [
            ("Titles",            ", ".join(pq.titles)),
            ("Variations",        ", ".join(pq.title_variations)),
            ("Locations",         ", ".join(pq.locations)),
            ("Remote OK",         str(pq.remote_ok)),
            ("Required Skills",   ", ".join(pq.skills_required)),
            ("Excluded Keywords", ", ".join(pq.exclude_keywords)),
        ]:
            ws.cell(row=r, column=1, value=f"  {label}").font = Font(italic=True)
            ws.cell(row=r, column=2, value=value)
            r += 1

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 80
    for row_num in range(1, r):
        ws.row_dimensions[row_num].height = 20
